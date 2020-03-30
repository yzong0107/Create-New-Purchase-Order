__author__ = "Tim Zong (yzong@ualberta.ca)"

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW
import time
import getpass
import openpyxl
from datetime import datetime

# file_loc = "..\excel file\download.xlsx"
# wb = openpyxl.load_workbook(file_loc)
# ws = wb.worksheets[0]
# print (ws.max_column,type(ws.max_column))
# ws.cell(1,1).fill = PatternFill(fgColor=YELLOW, fill_type = "solid")
# wb.save(file_loc)
print ("Tim Zong".split(" "))