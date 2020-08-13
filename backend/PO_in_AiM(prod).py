__author__ = "Tim Zong (yzong@ualberta.ca)"

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
import time
import getpass
import traceback
import openpyxl
from datetime import datetime
import glob
import numpy as np

class PurchaseOrder():
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.vars = {}
        # TODO: remember to update demo to prod
        self.instance="aimprod"

    def teardown_method(self):
        self.driver.quit()

    def login(self):
        url = "https://www."+self.instance+".ualberta.ca/fmax/screen/WORKDESK"
        self.driver.get(url)
        self.driver.set_window_size(1900, 1020)
        username = input('Enter your username: ')
        password = getpass.getpass('Enter your password : ')
        self.driver.find_element(By.ID, "username").send_keys(username)
        self.driver.find_element(By.ID, "password").send_keys(password)
        self.driver.find_element(By.ID, "login").click()
        self.driver.find_element(By.ID, "mainForm:menuListMain:PURCHASING").click()

    def search_WO(self,WO,phase):
        self.driver.find_element(By.ID, "mainForm:headerInclude:aimTitle1").click()
        self.driver.find_element(By.ID, "mainForm:menuListMain:WORKMGT").click()
        self.driver.find_element(By.ID, "mainForm:menuListMain:search_WO_VIEW").click()
        self.driver.find_element(By.ID, "mainForm:buttonPanel:reset").click()
        self.driver.find_element(By.ID, "mainForm:ae_p_pro_e_proposal").click()
        self.driver.find_element(By.ID, "mainForm:ae_p_pro_e_proposal").send_keys(WO)
        self.driver.find_element(By.ID, "mainForm:ae_p_phs_e_sort_code").click()
        self.driver.find_element(By.ID, "mainForm:ae_p_phs_e_sort_code").send_keys(phase)
        self.driver.find_element(By.ID, "mainForm:buttonPanel:executeSearch").click()
        try:
            self.driver.find_element(By.ID, "mainForm:browse:0:ae_p_pro_e_proposal").click()
            xpath ="//a[contains(text(),\'"+phase+"\')]"
            self.driver.find_element(By.XPATH, xpath).click()
            """save four lines below, in case xpath crashes"""
            # mytable = self.driver.find_element(By.ID,"mainForm:WO_VIEW_content:oldPhaseList")
            # for row in mytable.find_elements_by_css_selector('tr'):
            #     for cell in row.find_elements_by_tag_name('td'):
            #         print(cell.text)
            cppo = self.driver.find_element(By.ID, "mainForm:PHASE_VIEW_content:cpCompZoom:cpZoom0").text
            self.driver.find_element(By.ID, "mainForm:headerInclude:aimTitle1").click()
            self.driver.find_element(By.ID, "mainForm:menuListMain:PURCHASING").click()
            if cppo.strip()!="":
                return cppo
            else:
                return None
        except NoSuchElementException:
            self.driver.find_element(By.ID, "mainForm:headerInclude:aimTitle1").click()
            self.driver.find_element(By.ID, "mainForm:menuListMain:PURCHASING").click()
            return None

    def search_PO(self,po_no):
        self.driver.find_element(By.ID,"mainForm:buttonPanel:reset").click()
        time.sleep(0.5)
        self.driver.find_element(By.ID, "mainForm:ae_i_poe_e_description").send_keys(po_no)
        self.driver.find_element(By.ID, "mainForm:buttonPanel:executeSearch").click()
        try:
            self.driver.find_element(By.ID, "mainForm:browse:0:ae_i_poe_e_po_code").click()
            return True
        except:
            return False

    def log_po(self,po_no,supplier_no,person,item,line_total,WO,phase,material,currency):
        try:
            if pd.isna(WO): WO=""
            else: WO=WO.strip()
            if pd.isna(phase): phase = ""
            else: phase=phase.strip()
            item = item.upper()  # convert description to upper case
            line_item = WO + " - " + phase +"\n"+ item
            full_name = person.split(" ",1)
            cppo = self.search_WO(WO,phase)
            self.driver.find_element(By.ID, "mainForm:menuListMain:search_PO_VIEW").click()
            if self.search_PO(po_no):
                return None, "PO number already exists in AiM"
            else:
                self.driver.find_element(By.ID, "mainForm:buttonPanel:new").click()

            """PO main page"""
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.ID, 'mainForm:PO_EDIT_content:ae_i_poe_e_description')))
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:ae_i_poe_e_description").click()
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:ae_i_poe_e_description").send_keys(item)
            WebDriverWait(self.driver,5).until(lambda driver:self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:ae_i_poe_e_description").get_attribute("value")==item)
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:contractorZoom:contractorZoom0").click()
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:contractorZoom:contractorZoom0").send_keys(supplier_no)
            WebDriverWait(self.driver,5).until(lambda driver:self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:contractorZoom:contractorZoom0").get_attribute("value")==supplier_no)
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:contractorZoom:contractorZoom1").send_keys("1")
            if currency=="USD":
                #updates May 13, 2020
                self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:termsZoom:termsZoom01").send_keys("3")
            else:
                self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:termsZoom:termsZoom01").send_keys("1")
            if cppo is None:
                self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusTypeZoom:level0").send_keys("e-pro")
                WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusTypeZoom:level0").get_attribute("value") == "e-pro")
            else:
                self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusTypeZoom:level0").send_keys("cppo")
                WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusTypeZoom:level0").get_attribute("value") == "cppo")

            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").click()
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").send_keys("open")
            WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").get_attribute("value") == "open")
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:defaultWoZoom:defaultWorkOrder").click()
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:defaultWoZoom:defaultWorkOrder").send_keys(WO)
            WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:defaultWoZoom:defaultWorkOrder").get_attribute("value") == WO)
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:defaultWoZoom:defaultPhase").click()
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:defaultWoZoom:defaultPhase").send_keys(phase)
            WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:defaultWoZoom:defaultPhase").get_attribute("value") == phase)
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:disbDefaultsLineItem").click()
            dropdown = self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:disbDefaultsLineItem")
            dropdown.find_element(By.XPATH, "//option[. = 'Service']").click()
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:disbDefaultsLineItem").click()

            if cppo is None:
                udf_id = "mainForm:sideButtonPanel:moreMenu_3"
                self.driver.find_element(By.CSS_SELECTOR, "#mainForm\\3APO_EDIT_content\\3AtermsZoom\\3AtermsZoom01_button > .halflings").click()
                time.sleep(0.5)
            else:
                udf_id = "mainForm:sideButtonPanel:moreMenu_4"
                self.driver.find_element(By.CSS_SELECTOR, "#mainForm\\3APO_EDIT_content\\3ApoStatusTypeZoom\\3Alevel0_button > .halflings").click()
                time.sleep(0.5)
                self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:projectContract:level0").send_keys(cppo)
                WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:projectContract:level0").get_attribute("value") == cppo)
                self.driver.find_element(By.CSS_SELECTOR, "#mainForm\\3APO_EDIT_content\\3AtermsZoom\\3AtermsZoom01_button > .halflings").click()
                time.sleep(0.5)
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:placedbyZoom:placedbyZoom0").clear()
            self.driver.find_element(By.CSS_SELECTOR, "#mainForm\\3APO_EDIT_content\\3AplacedbyZoom\\3AplacedbyZoom0_button > .halflings").click()
            time.sleep(0.5)
            self.driver.find_element(By.ID, "mainForm:buttonPanel:search").click()
            time.sleep(0.5)
            self.driver.find_element(By.ID, "mainForm:ae_h_emp_e_fname").send_keys(full_name[0])
            WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:ae_h_emp_e_fname").get_attribute("value") == full_name[0])
            self.driver.find_element(By.ID, "mainForm:ae_h_emp_e_lname").send_keys(full_name[-1])
            WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:ae_h_emp_e_lname").get_attribute("value") == full_name[-1])
            self.driver.find_element(By.ID, "mainForm:buttonPanel:executeSearch").click()
            self.driver.find_element(By.ID, "mainForm:zoomTable:0:ae_h_emp_e_shop_person").click()

            try:
                """Line item"""
                self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:oldPoLineItemsList:addLineItemButton").click()
            except NoSuchElementException:
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                error_message = "Supplier may not exists in AiM, please double check"
                return None, error_message
            self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:ae_i_poe_d_vend_dsc").click()
            self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:ae_i_poe_d_vend_dsc").send_keys(line_item)
            WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:ae_i_poe_d_vend_dsc").get_attribute("value")==line_item)
            self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:amountValueServices").clear()
            self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:amountValueServices").send_keys(line_total)
            WebDriverWait(self.driver, 5).until(
                lambda driver: self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:amountValueServices").get_attribute("value") == line_total)
            self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:subledgerValue").click()
            dropdown = self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:subledgerValue")
            if material.upper()=="MATERIAL":
                dropdown.find_element(By.XPATH, "//option[. = 'Material']").click()
            elif material.upper()=="CONTRACT":
                dropdown.find_element(By.XPATH, "//option[. = 'Contract']").click()
            else:
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                error_message = "Please type in 'Material' or 'Contract' to indicate this PO's subledger"
                return None,error_message
            self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:subledgerValue").click()
            self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:link").click()
            time.sleep(1)
            self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()
            try:#2020-04-30: when WO or phase is empty
                """UDF"""
                self.driver.find_element(By.ID, udf_id).click()
                time.sleep(0.5)
                self.driver.find_element(By.ID, "mainForm:PO_UDF_EDIT_content:ae_i_poe_e_udf_custom001").send_keys(po_no)
                WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:PO_UDF_EDIT_content:ae_i_poe_e_udf_custom001").get_attribute("value")==po_no)
                self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()
                time.sleep(0.5)
            except:
                self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()
                error_message = self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:messages").text
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                return None, error_message
            """Change status to Finalized"""
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").clear()
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").send_keys("finalized")
            WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").get_attribute("value")=="finalized")
            self.driver.find_element(By.ID, "mainForm:buttonPanel:save").click()
            try:
                self.driver.find_element(By.ID, "mainForm:buttonControls:yes").click()
            except NoSuchElementException:
                pass
            WebDriverWait(self.driver,5).until(EC.presence_of_element_located((By.ID,"mainForm:PO_VIEW_content:ae_i_poe_e_purchase_order")))
            aim_po = self.driver.find_element(By.ID, "mainForm:PO_VIEW_content:ae_i_poe_e_purchase_order").text
            return aim_po,None
        except:
            try:
                error_message = self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:messages").text
            except NoSuchElementException:
                self.driver.find_element(By.ID,"mainForm:buttonControls:yes").click()
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            return None,error_message

    def multiple_lines(self,item,WO,phase,line_total,material):
        """Change status back to open"""
        WO = WO.strip()
        phase = phase.strip()
        line_item = WO + " - " + phase + "\n" + item.upper() #update on May 13, 2020
        self.driver.find_element(By.ID, "mainForm:buttonPanel:edit").click()
        time.sleep(0.5)
        self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").clear()
        self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").send_keys("open")
        WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").get_attribute("value") == "open")
        self.driver.find_element(By.CSS_SELECTOR, "#mainForm\\3APO_EDIT_content\\3ApoStatusZoom\\3Alevel0_button > .halflings").click()
        time.sleep(0.5)
        """Add a new line"""
        self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:oldPoLineItemsList:addLineItemButton").click()
        self.driver.find_element(By.ID, "mainForm:PO_ADD_LINE_ITEM_content:inentorytype3").click()
        time.sleep(0.5)
        self.driver.find_element(By.ID, "mainForm:buttonPanel:zoomNext").click()
        time.sleep(0.5)
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:ae_i_poe_d_vend_dsc").click()
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:ae_i_poe_d_vend_dsc").send_keys(line_item)
        WebDriverWait(self.driver, 5).until(
            lambda driver: self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:ae_i_poe_d_vend_dsc").get_attribute("value") == line_item)
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:amountValueServices").clear()
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:amountValueServices").send_keys(line_total)
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:subledgerValue").click()
        dropdown = self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:subledgerValue")
        if material.upper() == "MATERIAL":
            dropdown.find_element(By.XPATH, "//option[. = 'Material']").click()
        elif material.upper() == "CONTRACT":
            dropdown.find_element(By.XPATH, "//option[. = 'Contract']").click()
        else:
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            error_message = "Please type in 'Material' or 'Contract' to indicate this PO's subledger"
            return None, error_message
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:subledgerValue").click()
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:oldPoDisburList:0:seqLink").click()
        time.sleep(0.5)
        try:
            cppo = self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_DISBUR_EDIT_content:cpValue").text
            WO_id = "mainForm:PO_LINE_ITEM_DISBUR_EDIT_content:wophaseForCPZoom:wophaseZoom0"
            phase_id = "mainForm:PO_LINE_ITEM_DISBUR_EDIT_content:wophaseForCPZoom:wophaseZoom1"
        except NoSuchElementException:
            WO_id = "mainForm:PO_LINE_ITEM_DISBUR_EDIT_content:wophaseZoom:wophaseZoom0"
            phase_id= "mainForm:PO_LINE_ITEM_DISBUR_EDIT_content:wophaseZoom:wophaseZoom1"
        self.driver.find_element(By.ID, WO_id).click()
        self.driver.find_element(By.ID, WO_id).send_keys(WO)
        WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, WO_id).get_attribute("value") == WO)
        self.driver.find_element(By.ID, phase_id).click()
        self.driver.find_element(By.ID, phase_id).send_keys(phase)
        WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, phase_id).get_attribute("value") == phase)
        self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()
        time.sleep(0.5)
        self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()
        time.sleep(0.5)

        """Change status to Finalized"""
        try:
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").clear()
        except NoSuchElementException:
            error_message = self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_DISBUR_EDIT_content:messages").text
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            return None, error_message
        time.sleep(0.5)
        self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").send_keys("finalized")
        WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").get_attribute("value") == "finalized")
        self.driver.find_element(By.ID, "mainForm:buttonPanel:save").click()
        try:
            self.driver.find_element(By.ID, "mainForm:buttonControls:yes").click()
        except NoSuchElementException:
            pass
        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "mainForm:PO_VIEW_content:ae_i_poe_e_purchase_order")))
        aim_po = self.driver.find_element(By.ID, "mainForm:PO_VIEW_content:ae_i_poe_e_purchase_order").text
        return aim_po, None

    def search_cp(self,po_no,contr_admin):
        self.driver.find_element(By.ID, "mainForm:buttonPanel:reset").click()
        time.sleep(0.5)
        if contr_admin.upper()=="CONSTRUCTION":
            self.driver.find_element(By.ID, "mainForm:ae_cp_construct_con_e_description").click()
            self.driver.find_element(By.ID, "mainForm:ae_cp_construct_con_e_description").send_keys(po_no)
            self.driver.find_element(By.ID, "mainForm:buttonPanel:executeSearch").click()
            try:
                self.driver.find_element(By.ID, "mainForm:browse:0:ae_cp_construct_con_e_contract_no").click()
                return True
            except NoSuchElementException:
                return False
        elif contr_admin.upper()=="CONSULTANT":
            self.driver.find_element(By.ID, "mainForm:ae_cp_consult_con_e_description").click()
            self.driver.find_element(By.ID, "mainForm:ae_cp_consult_con_e_description").send_keys(po_no)
            self.driver.find_element(By.ID, "mainForm:buttonPanel:executeSearch").click()
            try:
                self.driver.find_element(By.ID, "mainForm:browse:0:ae_cp_consult_con_e_contract_no").click()
                return True
            except NoSuchElementException:
                return False

    def log_cp_construction(self,po_no,supplier,supplier_no,item,line_total,cp,comp_gr,comp,type,currency):
        description = supplier + "\n" + item + "\n" + po_no + "\n"
        try:
            self.driver.find_element(By.ID, "mainForm:buttonPanel:new").click()
            self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:ae_cp_construct_con_e_description").click()
            self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:ae_cp_construct_con_e_description").send_keys(description)
            self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:constructionContractTypeZoom:constructionContractTypeZoomLevel0").click()
            self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:constructionContractTypeZoom:constructionContractTypeZoomLevel0").send_keys(type)
            self.driver.find_element(By.CSS_SELECTOR,
                                     "#mainForm\\3A CONSTRUCTION_CONTRACT_EDIT_content\\3A constructionContractTypeZoom\\3A constructionContractTypeZoomLevel0_button > .halflings").click()
            time.sleep(0.5)
            try:
                self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:refNo").click()
            except:
                error_message = "Please provide a valid type"
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                return None, error_message
            self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:refNo").send_keys(po_no)
            self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:capitalProjectZoom:capitalProjectZoomLevel0").click()
            self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:capitalProjectZoom:capitalProjectZoomLevel0").send_keys(cp)
            self.driver.find_element(By.CSS_SELECTOR,
                                     "#mainForm\\3A CONSTRUCTION_CONTRACT_EDIT_content\\3A capitalProjectZoom\\3A capitalProjectZoomLevel0_button > .halflings").click()
            time.sleep(0.5)
            cp_error_url = "https://www." + self.instance + ".ualberta.ca/fmax/screen/ZOOM_CAPITAL_PROJECT"
            if self.driver.current_url == cp_error_url:
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                return None, "CP# is not valid in AiM system"

            self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:contractorZoom:level0").click()
            self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:contractorZoom:level0").send_keys(supplier_no)
            self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:contractorZoom:level1").click()
            self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:contractorZoom:level1").send_keys("1")
            if currency=="CAD":
                self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:termsZoomFc:level0").click()
                self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:termsZoomFc:level0").send_keys("1")
            else:
                self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:termsZoomFc:level0").click()
                self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:termsZoomFc:level0").send_keys("3")
            self.driver.find_element(By.CSS_SELECTOR, "#mainForm\\3A CONSTRUCTION_CONTRACT_EDIT_content\\3AtermsZoomFc\\3Alevel0_button > .halflings").click()
            time.sleep(0.5)
            self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:oldConstructionContractLineItemsList:addLineItemButton").click()
            time.sleep(0.5)
            self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:projCompZoom:projCompZoomLevel0").click()
            try:
                self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:projCompZoom:projCompZoomLevel0").send_keys(comp_gr)
                self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:projCompZoom:projCompZoomLevel1").click()
                self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:projCompZoom:projCompZoomLevel1").send_keys(comp)
                time.sleep(0.5)
                self.driver.find_element(By.CSS_SELECTOR,
                                         "#mainForm\\3A CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content\\3AprojCompZoom\\3AprojCompZoomLevel0_button > .halflings").click()
                time.sleep(0.5)
                comp_error_url="https://www."+self.instance+".ualberta.ca/fmax/screen/ZOOM_PROJECT_COMPONENT_NO_TIME"
                if self.driver.current_url==comp_error_url:
                    self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                    self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                    self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                    return None, "Component group/component is not valid in AiM system"

                self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:baseAmt").click()
                self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:baseAmt").send_keys(line_total)
                self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:link").click()
                time.sleep(0.5)
                self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:save").click()
            except:
                error_message = self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:messages").text
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                return None,error_message
        except:
            error_message = self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:messages").text
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            return None, error_message

        try:
            error_message = self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:messages").text
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            return None, error_message
        except:
            WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.ID, "mainForm:CONSTRUCTION_CONTRACT_VIEW_content:ae_cp_construct_con_e_contract_no")))
            construction_id = self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_VIEW_content:ae_cp_construct_con_e_contract_no").text
            return construction_id,None

    def log_cp_consultant(self,po_no,supplier,supplier_no,item,line_total,cp,comp_gr,comp,type,currency):
        description = supplier + "\n" + item + "\n" + po_no + "\n"
        try:
            self.driver.find_element(By.ID, "mainForm:buttonPanel:new").click()
            self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:ae_cp_consult_con_e_description").click()
            self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:ae_cp_consult_con_e_description").send_keys(description)
            self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:consultingContractTypeZoom:scTypeZoomLevel0").click()
            self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:consultingContractTypeZoom:scTypeZoomLevel0").send_keys(type)
            self.driver.find_element(By.CSS_SELECTOR, "#mainForm\\3A CONSULTING_CONTRACT_EDIT_content\\3A consultingContractTypeZoom\\3AscTypeZoomLevel0_button > .halflings").click()
            time.sleep(0.5)
            try:
                self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:refNo").click()
            except:
                error_message = "Please provide a valid type"
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                return None, error_message
            self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:refNo").send_keys(po_no)
            self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:capitalProjectZoom:capitalProjectZoomLevel0").click()
            self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:capitalProjectZoom:capitalProjectZoomLevel0").send_keys(cp)
            self.driver.find_element(By.CSS_SELECTOR,
                                     "#mainForm\\3A CONSULTING_CONTRACT_EDIT_content\\3A capitalProjectZoom\\3A capitalProjectZoomLevel0_button > .halflings").click()
            time.sleep(0.5)
            cp_error_url="https://www."+self.instance+".ualberta.ca/fmax/screen/ZOOM_CAPITAL_PROJECT"
            if self.driver.current_url==cp_error_url:
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                return None, "CP# is not valid in AiM system"

            self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:contractorZoom:level0").click()
            self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:contractorZoom:level0").send_keys(supplier_no)
            self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:contractorZoom:level1").click()
            self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:contractorZoom:level1").send_keys("1")
            if currency=="CAD":
                self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:termsZoomFc:level0").click()
                self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:termsZoomFc:level0").send_keys("1")
            else:
                self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:termsZoomFc:level0").click()
                self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:termsZoomFc:level0").send_keys("3")
            self.driver.find_element(By.CSS_SELECTOR, "#mainForm\\3A CONSULTING_CONTRACT_EDIT_content\\3AtermsZoomFc\\3Alevel0_button > .halflings").click()
            time.sleep(0.5)
            self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:oldSCLineItemsList:addLineItemButton").click()
            time.sleep(0.5)
            self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:projCompZoom:projCompZoomLevel1").click()
            try:
                self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:projCompZoom:projCompZoomLevel1").send_keys(comp_gr)
                self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:projCompZoom:projCompZoomLevel2").click()
                self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:projCompZoom:projCompZoomLevel2").send_keys(comp)
                time.sleep(0.5)
                self.driver.find_element(By.CSS_SELECTOR, "#mainForm\\3A CONSULTING_CONTRACT_DETAIL_EDIT_content\\3AprojCompZoom\\3AprojCompZoomLevel1_button > .halflings").click()
                time.sleep(0.5)
                comp_error_url = "https://www." + self.instance + ".ualberta.ca/fmax/screen/ZOOM_PROJECT_COMPONENT_NO_TIME"
                if self.driver.current_url == comp_error_url:
                    self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                    self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                    self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                    return None, "Component group/component is not valid in AiM system"

                self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:awardAmt").click()
                self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:awardAmt").send_keys(line_total)
                self.driver.find_element(By.CSS_SELECTOR, ".refreshLink").click()
                time.sleep(0.5)
                self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:save").click()
            except:
                error_message = self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:messages").text
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                return None, error_message
        except:
            error_message = self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:messages").text
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            return None, error_message

        try:
            error_message = self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:messages").text
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            return None, error_message
        except:
            WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.ID, "mainForm:CONSULTING_CONTRACT_VIEW_content:ae_cp_consult_con_e_contract_no")))
            consult_id = self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_VIEW_content:ae_cp_consult_con_e_contract_no").text
            return consult_id, None

    def log_cp(self,po_no,supplier,supplier_no,item,line_total,cp,comp_gr,comp,contr_admin,type,currency):
        contr_admin = contr_admin.strip()
        self.driver.find_element(By.ID, "mainForm:headerInclude:aimTitle1").click()
        self.driver.find_element(By.ID, "mainForm:menuListMain:CONTRACT").click()
        if contr_admin.upper()=="CONSTRUCTION":
            self.driver.find_element(By.ID, "mainForm:menuListMain:search_CONSTRUCTION_CONTRACT_VIEW").click()
        elif contr_admin.upper()=="CONSULTANT":
            self.driver.find_element(By.ID, "mainForm:menuListMain:search_CONSULTING_CONTRACT_VIEW").click()
        else:
            self.driver.find_element(By.ID,"mainForm:headerInclude:aimTitle1").click()
            error_message = "Please type in 'construction' or 'consultant' to indicate this CPPM's type"
            return None, error_message #skip the line

        """search for the cp first"""
        if self.search_cp(po_no,contr_admin):
            error_message = "PO# {0} of CPPM {1} is already in the system as type of {2}".format(po_no,cp,contr_admin)
            return None, error_message

        """create a new record"""
        if contr_admin.upper()=="CONSTRUCTION":
            saved,error = self.log_cp_construction(po_no,supplier,supplier_no,item,line_total,cp,comp_gr,comp,type,currency)
        else: #consultant
            saved, error = self.log_cp_consultant(po_no,supplier,supplier_no,item,line_total,cp,comp_gr,comp,type,currency)

        return saved,error

    def multiple_construction_lines(self,comp_gr,comp,line_total):
        self.driver.find_element(By.ID, "mainForm:buttonPanel:edit").click()
        time.sleep(0.5)
        self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:oldConstructionContractLineItemsList:addLineItemButton").click()
        self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:projCompZoom:projCompZoomLevel0").click()
        self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:projCompZoom:projCompZoomLevel0").send_keys(comp_gr)
        self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:projCompZoom:projCompZoomLevel1").click()
        self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:projCompZoom:projCompZoomLevel1").send_keys(comp)
        self.driver.find_element(By.CSS_SELECTOR,
                                 "#mainForm\\3A CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content\\3AprojCompZoom\\3AprojCompZoomLevel1_button > .halflings").click()
        time.sleep(0.5)
        comp_error_url = "https://www." + self.instance + ".ualberta.ca/fmax/screen/ZOOM_PROJECT_COMPONENT_NO_TIME"
        if self.driver.current_url == comp_error_url:
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            return None, "Component group/component is not valid in AiM system"

        self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:baseAmt").click()
        self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:baseAmt").send_keys(line_total)
        self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_LINE_ITEM_EDIT_content:link").click()
        time.sleep(0.5)
        self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()
        self.driver.find_element(By.ID, "mainForm:buttonPanel:save").click()

        try:
            error_message = self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_EDIT_content:messages").text
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            return None, error_message
        except:
            WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.ID, "mainForm:CONSTRUCTION_CONTRACT_VIEW_content:ae_cp_construct_con_e_contract_no")))
            construction_id = self.driver.find_element(By.ID, "mainForm:CONSTRUCTION_CONTRACT_VIEW_content:ae_cp_construct_con_e_contract_no").text
            return construction_id, None

    def multiple_consultant_lines(self,comp_gr,comp,line_total):
        self.driver.find_element(By.ID, "mainForm:buttonPanel:edit").click()
        time.sleep(0.5)
        self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:oldSCLineItemsList:addLineItemButton").click()
        self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:projCompZoom:projCompZoomLevel1").click()
        self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:projCompZoom:projCompZoomLevel1").send_keys(comp_gr)
        self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:projCompZoom:projCompZoomLevel2").click()
        self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:projCompZoom:projCompZoomLevel2").send_keys(comp)
        self.driver.find_element(By.CSS_SELECTOR,
                                 "#mainForm\\3A CONSULTING_CONTRACT_DETAIL_EDIT_content\\3AprojCompZoom\\3AprojCompZoomLevel2_button > .halflings").click()
        time.sleep(0.5)
        comp_error_url = "https://www." + self.instance + ".ualberta.ca/fmax/screen/ZOOM_PROJECT_COMPONENT_NO_TIME"
        if self.driver.current_url == comp_error_url:
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            return None, "Component group/component is not valid in AiM system"

        self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:awardAmt").click()
        self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:awardAmt").send_keys(line_total)
        self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_DETAIL_EDIT_content:link").click()
        time.sleep(0.5)
        self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()
        self.driver.find_element(By.ID, "mainForm:buttonPanel:save").click()

        try:
            error_message = self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_EDIT_content:messages").text
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            return None, error_message
        except:
            WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.ID, "mainForm:CONSULTING_CONTRACT_VIEW_content:ae_cp_consult_con_e_contract_no")))
            consult_id = self.driver.find_element(By.ID, "mainForm:CONSULTING_CONTRACT_VIEW_content:ae_cp_consult_con_e_contract_no").text
            return consult_id, None

def write_to_log_title(file_location,col_num):
    wb = openpyxl.load_workbook(file_location)
    ws = wb.worksheets[0]
    # TODO: remember to update demo to prod
    ws.cell(row=1, column=col_num+1).value = "AiM PO(prod)"  # column N
    ws.cell(row=1, column=col_num+1).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
    ws.cell(row=1, column=col_num+2).value = "Time stamp(prod)"  # column O
    ws.cell(row=1, column=col_num+2).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
    ws.cell(row=1, column=col_num+3).value = "Error Messages(prod)"  # column P
    ws.cell(row=1, column=col_num+3).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
    wb.save(file_location)

def write_to_log(file_location,row,aim_po,error,col_num):
    wb = openpyxl.load_workbook(file_location)
    ws = wb.worksheets[0]
    if aim_po is not None:
        ws.cell(row=row+2, column=col_num+1).value = aim_po  # column N
        ws.cell(row=row+2, column=col_num+1).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
        ws.cell(row=row+2, column=col_num+2).value = datetime.now()  # column O
        ws.cell(row=row+2, column=col_num+2).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
        ws.cell(row=row+2, column=col_num+3).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
    elif error is not None:
        ws.cell(row=row + 2, column=col_num+3).value = error  # column P
        ws.cell(row=row + 2, column=col_num+3).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
        ws.cell(row=row + 2, column=col_num+2).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
        ws.cell(row=row + 2, column=col_num+1).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
    wb.save(file_location)



if __name__ == '__main__':
    file_loc = glob.glob('V:\Purchasing Astro Boy\commitment files\Input\*.xlsx')[0]  # assuming only 1 excel file in this folder

    new_po = PurchaseOrder()
    new_po.setup_method()
    new_po.login()

    start_time = time.time()
    sheet = pd.read_excel(file_loc, dtype=str)
    col_num = sheet.shape[1]
    write_to_log_title(file_loc,col_num)
    saved_PO=[]
    for i in range(sheet.shape[0]):
        saved_PO = list(set(saved_PO))
        po_no,supp, supplier_no,person, item, line_total, WO, phase,CP,comp_gr,comp,_,contr_admin,subleger = sheet.iloc[i,:14].values
        contr_admin = contr_admin.strip()
        currency,_ = sheet.iloc[i,14:16].values
        if pd.notna(CP):
            #handle the PO with CP number
            if i>0:
                if sheet.iloc[i,0]==sheet.iloc[i-1,0] and sheet.iloc[i,0] in saved_PO:
                    if contr_admin.upper()=="CONSTRUCTION":
                        cppm, error = new_po.multiple_construction_lines(comp_gr,comp,line_total)
                    else:
                        cppm, error = new_po.multiple_consultant_lines(comp_gr, comp, line_total)
                    write_to_log(file_loc, i, cppm, error, col_num)
                    print("row {} is processed, Contract id is : {}".format(i + 2, cppm))
                    continue
            cppm,error = new_po.log_cp(po_no,supp,supplier_no,item,line_total,CP,comp_gr,comp,contr_admin,subleger,currency)
            write_to_log(file_loc,i,cppm,error,col_num)
            if error is None:
                saved_PO.append(sheet.iloc[i,0])
            print ("row {} is processed, contract id is : {}".format(i+2,cppm))
            continue
        else:
            if i>0:
                if sheet.iloc[i,0]==sheet.iloc[i-1,0] and sheet.iloc[i,0] in saved_PO:#if this line has same PO number to the line above
                    aim_po, error = new_po.multiple_lines(item,WO,phase,line_total,subleger)
                    write_to_log(file_loc, i, aim_po, error,col_num)
                    print("row {} is processed, AiM PO is : {}".format(i + 2, aim_po))
                    continue
            aim_po,error = new_po.log_po(po_no, supplier_no,person, item, line_total, WO, phase,subleger,currency)
            write_to_log(file_loc,i,aim_po,error,col_num)
            if error is None:
                saved_PO.append(sheet.iloc[i,0])
            print ("row {} is processed, AiM PO is : {}".format(i+2,aim_po))
    time_taken = time.time()-start_time
    print("")
    print("***************************************")
    print("Done! Time taken: {:.2f}s ({:.2f}min)".format(time_taken, time_taken / 60.))
    print("Please go to excel file to double check")
    print("***************************************")

