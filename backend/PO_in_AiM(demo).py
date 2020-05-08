__author__ = "Tim Zong (yzong@ualberta.ca)"

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
import time
import getpass
import traceback
import openpyxl
from datetime import datetime
import glob
import numpy as np

class PurchaseOrder():
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.vars = {}

    def teardown_method(self):
        self.driver.quit()

    def login(self):
        self.driver.get("https://www.aimdemo.ualberta.ca/fmax/screen/WORKDESK")
        self.driver.set_window_size(1900, 1020)
        username = input('Enter your username: ')
        password = getpass.getpass('Enter your password : ')
        self.driver.find_element(By.ID, "username").send_keys(username)
        self.driver.find_element(By.ID, "password").send_keys(password)
        self.driver.find_element(By.ID, "login").click()
        self.driver.find_element(By.ID, "mainForm:menuListMain:PURCHASING").click()

    def log_po(self,po_no,supplier_no,person,item,line_total,WO,phase,material,first_PO=True):
        try:
            if pd.isna(WO): WO=""
            if pd.isna(phase): phase = ""
            line_item = WO + " - " + phase
            item = item.upper()  # convert description to upper case
            full_name = person.split(" ")
            if first_PO:
                self.driver.find_element(By.ID, "mainForm:menuListMain:new_PO_VIEW").click()
            else:
                self.driver.find_element(By.ID, "mainForm:buttonPanel:new").click()
            """PO main page"""
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.ID, 'mainForm:PO_EDIT_content:ae_i_poe_e_description')))
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:ae_i_poe_e_description").send_keys(item)
            WebDriverWait(self.driver,5).until(lambda driver:self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:ae_i_poe_e_description").get_attribute("value")==item)
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:contractorZoom:contractorZoom0").send_keys(supplier_no)
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:contractorZoom:contractorZoom1").send_keys("1")
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:termsZoom:termsZoom01").send_keys("1")
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusTypeZoom:level0").send_keys("e-pro")
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:placedbyZoom:placedbyZoom0").clear()
            self.driver.find_element(By.CSS_SELECTOR, "#mainForm\\3APO_EDIT_content\\3AplacedbyZoom\\3AplacedbyZoom0_button > .halflings").click()
            self.driver.find_element(By.ID, "mainForm:buttonPanel:search").click()
            self.driver.find_element(By.ID, "mainForm:ae_h_emp_e_fname").send_keys(full_name[0])
            self.driver.find_element(By.ID, "mainForm:ae_h_emp_e_lname").send_keys(full_name[-1])
            self.driver.find_element(By.ID, "mainForm:buttonPanel:executeSearch").click()
            self.driver.find_element(By.ID, "mainForm:zoomTable:0:ae_h_emp_e_shop_person").click()

            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").send_keys("open")
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:defaultWoZoom:defaultWorkOrder").send_keys(WO)
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:defaultWoZoom:defaultPhase").send_keys(phase)
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:disbDefaultsLineItem").click()
            dropdown = self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:disbDefaultsLineItem")
            dropdown.find_element(By.XPATH, "//option[. = 'Service']").click()
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:disbDefaultsLineItem").click()
            self.driver.find_element(By.CSS_SELECTOR, "#mainForm\\3APO_EDIT_content\\3AtermsZoom\\3AtermsZoom01_button > .halflings").click()
            time.sleep(0.5)
            try:
                """Line item"""
                self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:oldPoLineItemsList:addLineItemButton").click()
            except NoSuchElementException:
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                error_message = "Supplier may not exists in AiM, please double check"
                return None, error_message
            self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:ae_i_poe_d_vend_dsc").click()
            self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:ae_i_poe_d_vend_dsc").send_keys(line_item)
            WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:ae_i_poe_d_vend_dsc").get_attribute("value")==line_item)
            self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:amountValueServices").clear()
            self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:amountValueServices").send_keys(line_total)
            self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:subledgerValue").click()
            dropdown = self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:subledgerValue")
            if material.upper()=="MATERIAL":
                dropdown.find_element(By.XPATH, "//option[. = 'Material']").click()
            elif material.upper()=="CONTRACT":
                dropdown.find_element(By.XPATH, "//option[. = 'Contract']").click()
            else:
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                error_message = "Please type in 'Material' or 'Contract' to indicate this PO's subledger"
                return None,error_message
            self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:subledgerValue").click()
            self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()
            try:#2020-04-30: when WO or phase is empty
                """UDF"""
                self.driver.find_element(By.ID, "mainForm:sideButtonPanel:moreMenu_3").click()
                time.sleep(0.3)
                self.driver.find_element(By.ID, "mainForm:PO_UDF_EDIT_content:ae_i_poe_e_udf_custom001").send_keys(po_no)
                WebDriverWait(self.driver, 5).until(lambda driver: self.driver.find_element(By.ID, "mainForm:PO_UDF_EDIT_content:ae_i_poe_e_udf_custom001").get_attribute("value")==po_no)
                self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()
            except:
                self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()
                error_message = self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:messages").text
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
                return None, error_message
            """Change status to Finalized"""
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").clear()
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").send_keys("finalized")

            self.driver.find_element(By.ID, "mainForm:buttonPanel:save").click()
            aim_po = self.driver.find_element(By.ID, "mainForm:PO_VIEW_content:ae_i_poe_e_purchase_order").text
            return aim_po,None
        except NoSuchElementException:
            error_message = self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:messages").text
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            return None,error_message

    def multiple_lines(self,WO,phase,line_total,material):
        """Change status back to open"""
        line_item = WO + " - " + phase
        self.driver.find_element(By.ID, "mainForm:buttonPanel:edit").click()
        self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").clear()
        self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").send_keys("open")
        self.driver.find_element(By.CSS_SELECTOR, "#mainForm\\3APO_EDIT_content\\3ApoStatusZoom\\3Alevel0_button > .halflings").click()
        time.sleep(0.5)
        """Add a new line"""
        self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:oldPoLineItemsList:addLineItemButton").click()
        self.driver.find_element(By.ID, "mainForm:PO_ADD_LINE_ITEM_content:inentorytype3").click()
        time.sleep(0.2)
        self.driver.find_element(By.ID, "mainForm:buttonPanel:zoomNext").click()
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:ae_i_poe_d_vend_dsc").click()
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:ae_i_poe_d_vend_dsc").send_keys(line_item)
        WebDriverWait(self.driver, 5).until(
            lambda driver: self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:ae_i_poe_d_vend_dsc").get_attribute("value") == line_item)
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:amountValueServices").clear()
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:amountValueServices").send_keys(line_total)
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:subledgerValue").click()
        dropdown = self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:subledgerValue")
        if material.upper() == "MATERIAL":
            dropdown.find_element(By.XPATH, "//option[. = 'Material']").click()
        elif material.upper() == "CONTRACT":
            dropdown.find_element(By.XPATH, "//option[. = 'Contract']").click()
        else:
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            error_message = "Please type in 'Material' or 'Contract' to indicate this PO's subledger"
            return None, error_message
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:subledgerValue").click()
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_EDIT_content:oldPoDisburList:0:seqLink").click()
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_DISBUR_EDIT_content:wophaseZoom:wophaseZoom0").click()
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_DISBUR_EDIT_content:wophaseZoom:wophaseZoom0").send_keys(WO)
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_DISBUR_EDIT_content:wophaseZoom:wophaseZoom1").click()
        self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_DISBUR_EDIT_content:wophaseZoom:wophaseZoom1").send_keys(phase)
        self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()
        self.driver.find_element(By.ID, "mainForm:buttonPanel:done").click()

        """Change status to Finalized"""
        try:
            self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").clear()
        except NoSuchElementException:
            error_message = self.driver.find_element(By.ID, "mainForm:PO_LINE_ITEM_DISBUR_EDIT_content:messages").text
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            self.driver.find_element(By.ID, "mainForm:buttonPanel:cancel").click()
            return None, error_message
        self.driver.find_element(By.ID, "mainForm:PO_EDIT_content:poStatusZoom:level0").send_keys("finalized")
        self.driver.find_element(By.ID, "mainForm:buttonPanel:save").click()
        aim_po = self.driver.find_element(By.ID, "mainForm:PO_VIEW_content:ae_i_poe_e_purchase_order").text
        return aim_po, None


def write_to_log_title(file_location):
    wb = openpyxl.load_workbook(file_location)
    ws = wb.worksheets[0]
    ws.cell(row=1, column=12).value = "AiM PO"  # column L
    ws.cell(row=1, column=12).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
    ws.cell(row=1, column=13).value = "Time stamp"  # column M
    ws.cell(row=1, column=13).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
    ws.cell(row=1, column=14).value = "Error Messages"  # column M
    ws.cell(row=1, column=14).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
    wb.save(file_location)

def write_to_log(file_location,row,aim_po,error):
    wb = openpyxl.load_workbook(file_location)
    ws = wb.worksheets[0]
    if aim_po is not None:
        ws.cell(row=row+2, column=12).value = aim_po  # column L
        ws.cell(row=row+2, column=12).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
        ws.cell(row=row+2, column=13).value = datetime.now()  # column M
        ws.cell(row=row+2, column=13).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
        ws.cell(row=row+2, column=14).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
    elif error is not None:
        ws.cell(row=row + 2, column=14).value = error  # column M
        ws.cell(row=row + 2, column=14).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
        ws.cell(row=row + 2, column=13).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
        ws.cell(row=row + 2, column=12).fill = PatternFill(fgColor=YELLOW, fill_type="solid")
    wb.save(file_location)



if __name__ == '__main__':
    file_loc = glob.glob('V:\Purchasing Astro Boy\commitment files\Input\*.xlsx')[0]  # assuming only 1 excel file in this folder
    write_to_log_title(file_loc)


    new_po = PurchaseOrder()
    new_po.setup_method()
    new_po.login()

    start_time = time.time()
    sheet = pd.read_excel(file_loc, dtype=str)
    first_po = True
    saved_PO=[]
    for i in range(sheet.shape[0]):
        saved_PO = list(set(saved_PO))
        po_no,_, supplier_no,person, item, line_total, WO, phase,CP,_,material = sheet.iloc[i,:11].values
        if pd.notna(CP):
            #TODO: handle the PO with CP number
            print ("row {} is NOT processed, as CP is not null".format(i+2))
            continue
        if i>0:
            if sheet.iloc[i,0]==sheet.iloc[i-1,0] and sheet.iloc[i,0] in saved_PO:#if this line has same PO number to the line above
                aim_po, error = new_po.multiple_lines(WO,phase,line_total,material)
                write_to_log(file_loc, i, aim_po, error)
                print("row {} is processed, AiM PO is : {}".format(i + 2, aim_po))
                continue
        aim_po,error = new_po.log_po(po_no, supplier_no,person, item, line_total, WO, phase,material,first_PO=first_po)
        write_to_log(file_loc,i,aim_po,error)
        if error is None:
            saved_PO.append(sheet.iloc[i,0])
        first_po = False
        print ("row {} is processed, AiM PO is : {}".format(i+2,aim_po))
    time_taken = time.time()-start_time
    print("")
    print("***************************************")
    print("Done! Time taken: {:.2f}s ({:.2f}min)".format(time_taken, time_taken / 60.))
    print("Please go to excel file to double check")
    print("***************************************")


